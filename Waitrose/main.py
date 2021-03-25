"""
Created on Thu Jan  7 09:37:59 2021
@author: kathrynhewitt
python3

main.py is the main python script 
extract information from waitrose.com and add information to a new excel

requirements 
    #pip install requests
    #pip install beautifulsoup4
"""

import framework 
import requests 
from bs4 import BeautifulSoup 
import openpyxl
import pandas as pd 
from datetime import date
from openpyxl import workbook
from datetime import date
import time

productNames = []
prices = []
amounts = []
pricePers = []
urls = []
divisionList = []
myDefinition = []

#open excel 
#extract each urls from excel and add to a list 
wb_obj = openpyxl.load_workbook('Price_comparison.xlsx')
sheet_obj = wb_obj.active
for i in range (4, 74):
    cell_obj = sheet_obj.cell(row=i, column=22)
    textd = cell_obj.value
    urls.append(textd)
#print(urls)
#count number in list to check all have been retrieved #l = len(urls) #print(l)
 
for i in range(4,74):
    cell_obj = sheet_obj.cell(row=i, column=23)
    myDef = cell_obj.value
    myDefinition.append(myDef)
     

for url in urls:
    getInformation()

division()
createExcel()





