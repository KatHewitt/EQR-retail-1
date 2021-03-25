"""
Created on Thu Jan  7 09:37:59 2021
@author: kathrynhewitt
python3

framework containing all the functions for the webscraper
functions will be called in the main.py
"""
import requests 
from bs4 import BeautifulSoup
from datetime import date
from openpyxl import Workbook
import time

start = "£"
end = "p"
end2 = "est."
s = "s"
g = "g"
h = "h"
kg = "kg"
minn = "min "
minn2 = "minimum"
typicalWeight = "Typical weight"   
each = "each"
og = "0g"
re = "re"
ml = "ml"
d = "d"
x = "x"
dot = "."

def getInformation():
    if url == None:
        print('none')
        productNames.append('none')
        prices.append(0)
        amounts.append(0)
        pricePers.append(0)
    else:
        result = requests.get(url)
        src = result.content
        soup = BeautifulSoup(src, 'lxml')
        
        for productName in soup.find_all('span',{'class':'name___30fwb'}):
            productNames.append(productName.text.strip())
            #print(productNames)
    
        for price in soup.find_all('span',{'data-test':'product-pod-price'}):
           # prices.append(price.text.strip())
            #print(prices)
            newPrice = price.text.strip()
            
            if end2 in newPrice  and newPrice[0] in start:
                splitPrice = newPrice.split('each est.')[0]
                splitPrice2 = splitPrice.split('£')[1]
                inte = float(splitPrice2)
                finalPrice = inte
                prices.append(finalPrice)
            
            elif newPrice[0] in start:
                splitPrice = newPrice.split('£')[1]
                inte = float(splitPrice)
                finalPrice = inte
                #print(splitPrice)
                prices.append(finalPrice)
           
            elif newPrice[-1] in end:
                    splitPrice = newPrice.split('p')[0]
                    intsplitPrice = float(splitPrice)
                    finalPrice = (intsplitPrice/100)
                    #print(finalPrice)
                    prices.append(finalPrice)
            elif newPrice[-1] in end2:
                    splitPrice = newPrice.split('p')[0]
                    intsplitPrice = float(splitPrice)
                    finalPrice = (intsplitPrice/100)
                    #print(finalPrice)
                    prices.append(finalPrice)
            else:
                #print(newPrice)
                prices.append(0)
        #print(prices)
    
        for amount in soup.find_all('span', {'class':'size___2HSwr sizeMessage___3o5Ri'}):
            newAmount = amount.text.strip()
            if newAmount[-1] in s and newAmount[0] in minn:
                splitAmount = newAmount.split('s')[0]
                splitAmount2 = splitAmount.split('min ')[1]
                inte = float(splitAmount2)
                finalAmount = inte
                #print(splitAmount2)
                amounts.append(finalAmount)
            #5s    
            elif newAmount[-1] in s:
                splitAmount = newAmount.split('s')[0]
                inte = float(splitAmount)
                finalAmount = inte
                #print(splitAmount)
                amounts.append(finalAmount)
            #each 
            elif newAmount == each:
                newAmount = 1
                #print(newAmount)
                amounts.append(newAmount)
            #typical weight 10kg
            elif newAmount[0] in typicalWeight and newAmount[-2] in kg:
                splitAmount = newAmount.split('Typical weight ')[1]
                splitAmount2 = splitAmount.split('kg')[0]
                inte = float(splitAmount2)
                finalAmount = inte
                #print(splitAmount2)
                amounts.append(finalAmount)
            #typical weight 100g
            elif newAmount[0] in typicalWeight and newAmount[-1] in g:
                splitAmount = newAmount.split('Typical weight ')[1]
                splitAmount2 = splitAmount.split('g')[0]
                intsplitAmount = float(splitAmount2)
                finalAmount = (intsplitAmount/1000)
                #print(finalAmount)
                amounts.append(finalAmount)
            #1kg
            elif newAmount[-2] in kg:
                splitAmount = newAmount.split('kg')[0]
                inte = float(splitAmount)
                finalAmount = inte
                #print(splitAmount)
                amounts.append(finalAmount)
            #min 6
            elif newAmount[0] in minn:
                splitAmount = newAmount[-1:]
                inte = float(splitAmount)
                finalAmount = inte
                #print(splitAmount)
                amounts.append(finalAmount)
            #minimum 6
            elif newAmount[0] in minn2:
                splitAmount = newAmount[-1:]
                inte = float(splitAmount)
                finalAmount = inte
                #print(splitAmount)
                amounts.append(finalAmount)
            #750ml
            elif newAmount[-2] in ml:
                splitAmount = newAmount.split('ml')[0]
                inte = float(splitAmount)
                finalAmount = (inte/1000)
                amounts.append(finalAmount)
            #2.2litre
            elif newAmount[-2] in re:
                splitAmount = newAmount.split('litre')[0]
                inte = float(splitAmount)
                finalAmount = inte
                amounts.append(finalAmount)
            #drained 4x150g
            elif newAmount[0] in d and newAmount[-1] in g:
                splitAmount = newAmount.split('drained ')[1]
                splitAmount2 = splitAmount.split('x')[0]
                inte = float(splitAmount2)
                finalAmount = inte
                amounts.append(finalAmount)
            #4x400g
            elif x in newAmount and newAmount[-1] in g:
                splitAmount = newAmount.split('x')[0]
                inte = float(splitAmount)
                finalAmount = inte
                amounts.append(finalAmount)
            #100g    
            elif newAmount[-1] in g:
                    splitAmount = newAmount.split('g')[0]
                    intsplitAmount = float(splitAmount)
                    finalAmount = (intsplitAmount/1000)
                    #print(finalAmount)
                    amounts.append(finalAmount) 
            #all else
            else:
                amounts.append(0)            
            #print(amounts)            
        
#        for pricePer in soup.find_all('span',{'class':'pricePerUnit___1L8TG'}):
#            newPriceP = pricePer.text.strip()
#            newPriceP2 = newPriceP[1:-1]
#            #50p each
#            if newPriceP2[-1] in h:
#                splitPriceP = newPriceP2.split('p each')[0]
#                inte = float(splitPriceP)
#                finalPriceP = (inte/100)
#                #print(finalPriceP)
#                pricePers.append(finalPriceP)
#            #£1/kg
#            elif newPriceP2[0] in start and newPriceP2[-2] in kg:
#                splitPriceP = newPriceP2.split('£')[1]
#                splitPriceP2 = splitPriceP.split('/kg')[0]
#                inte = float(splitPriceP2)
#                finalPriceP = inte
#                #print(splitPriceP2)
#                pricePers.append(finalPriceP)
#            #£1/100g    
#            elif newPriceP2[0] in start and newPriceP2[-2] in og:
#                splitPriceP = newPriceP2.split('£')[1]
#                splitPriceP2 = splitPriceP.split('/100g')[0]          
#                inte = float(splitPriceP2)
#                finalPriceP = (inte*10)
#                #print(finalPriceP)
#                pricePers.append(finalPriceP)
#            #50p/100g       
#            elif newPriceP2[-2] in og:
#                splitPriceP = newPriceP2.split('p/100g')[0]
#                inte = float(splitPriceP)
#                finalPriceP = (inte/10)
#                #print(finalPriceP)
#                pricePers.append(finalPriceP)
#            #50p/kg   
#            elif newPriceP2[-2] in kg:
#                splitPriceP = newPriceP2.split('p/kg')[0]
#                inte = float(splitPriceP)
#                finalPriceP = (inte/100)
#                #print(finalPriceP)
#                pricePers.append(finalPriceP)
#            #20p/litre
#            elif newPriceP2[-2] in re:
#                splitPriceP2 = newPriceP2.split('p/litre')[0]
#                inte = float(splitPriceP2)
#                finalPriceP = (inte/100)
#                pricePers.append(finalPriceP)
#            #20p/100ml
#            elif newPriceP2[-2] in ml:
#                splitPriceP2 = newPriceP2.split('p/100ml')[0]
#                inte = float(splitPriceP2)
#                finalPriceP = (inte/10)
#                pricePers.append(finalPriceP)
#            #else if not found
#            else:
#                #print("null") 
#                pricePers.append(0)
#            #print(pricePers)
#            time.sleep(1)

def division():
    for i, j in zip(prices, amounts):
        try:
            divisionList.append(i/j)
        except ZeroDivisionError:
            divisionList.append(0)


def createExcel():
    book = Workbook()
    sheet = book.active
    sheet['B1'] = "M&S on Ocado"
    sheet['J1'] = "Waitrose.com"
    rows = (

            (0,'Product Link','Product (my definition)','Website product name', 'Price',
             'Number per pack/weight/size', 'Price per item/kg/Litre', '','',
             'Product Link','Product (my definition)','Website product name', 'Price',
             'Number per pack/weight/size', 'Price per item/kg/Litre')
            )
    for i in range(1,15):
        cellref=sheet.cell(row=2,column=i)
        cellref.value=rows[i]  
    
    #column numbers 
    urlsCol = 9
    mydefCol = 10
    nameCol = 11
    priceCol = 12
    amountCol = 13
    pricePerCol = 14

    #writes urls to excel
    for i in range(len(urls)):
        cellref = sheet.cell(row=i+3, column=urlsCol)
        cellref.value = urls[i]
    #writes my definition to excel
    for i in range(len(myDefinition)):
        cellref = sheet.cell(row=i+3, column=mydefCol)
        cellref.value = myDefinition[i]
    #writes product names to excel
    for i in range(len(productNames)):
        cellref = sheet.cell(row=i+3, column=nameCol)
        cellref.value = productNames[i]
    #writes prices to excel
    for i in range(len(prices)):
        cellref = sheet.cell(row=i+3, column=priceCol)
        cellref.value = prices[i]
    #writes amounts to excel
    for i in range(len(amounts)):
        cellref = sheet.cell(row=i+3, column=amountCol)
        cellref.value = amounts[i]
     #writes price per to excel   
    for i in range(len(divisionList)):
        cellref = sheet.cell(row=i+3, column=pricePerCol)
        cellref.value = divisionList[i]
    
    #Find current date and save excel
    today = date.today()
    d4= today.strftime("%b-%d-%Y") 
    book.save(d4+'_price_comparison.xlsx')



    

