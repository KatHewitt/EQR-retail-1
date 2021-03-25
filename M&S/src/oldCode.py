import openpyxl
import time
import requests
from bs4 import BeautifulSoup

from src.framework import getProductName, getProductPrice, getProductAmount

urls = []
myDefinition = []

productNames = []
productPrices = []
productAmounts = []

# open price comparison excel
wb_obj = openpyxl.load_workbook('Price_comparison.xlsx')
sheet_obj = wb_obj.active
# extract each urls from excel and add to a list
for i in range(4, 74):
    cell_obj = sheet_obj.cell(row=i, column=8)
    textd = cell_obj.value
    urls.append(textd)
    urlLen = len(urls)
# extract each my definition from excel and add to a list
for i in range(4, 74):
    cell_obj = sheet_obj.cell(row=i, column=9)
    myDef = cell_obj.value
    myDefinition.append(myDef)
    mydefLen = len(myDefinition)

# tests for matching lenghts
#testLenUrlvMyDef()

for url in urls:
    result = requests.get(url)
    src = result.content
    soup = BeautifulSoup(src, 'lxml')
    try:
        name = soup.find('h2', {'class': ''}).next_element
        #print(name)
        productNames.append(name)
    except:
        productNames.append("None")
        print(url)
    try:
        price = soup.find('h2', {'class': 'bop-price__current'})
        #print(price.text.strip())
        productPrices.append(price.text.strip())
    except:
        productPrices.append(0)
        print(url)
    try:
        amount = soup.find('span', {'class': 'bop-catchWeight'})
        #print(amount.text.strip())
        productAmounts.append(amount.text.strip())
    except:
        productAmounts.append(0)
        print(url)
    time.sleep(5)

print(productNames)
print(productPrices)
print(productAmounts)

print(len(productNames))
print(len(productPrices))
print(len(productAmounts))