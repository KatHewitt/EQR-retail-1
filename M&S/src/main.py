import openpyxl
import time
import requests
from bs4 import BeautifulSoup

from src.framework import getProductName, getProductPrice, getProductAmount

urls = []
myDefinition = []

productNames = []
productPrices = []
productAmounts = []

attempts = 0
attempts2 = 0
attempts3 = 0

# open price comparison excel
wb_obj = openpyxl.load_workbook('Price_comparison.xlsx')
sheet_obj = wb_obj.active
# extract each urls from excel and add to a list
for i in range(4, 74):
    cell_obj = sheet_obj.cell(row=i, column=8)
    textd = cell_obj.value
    urls.append(textd)
    urlLen = len(urls)
# extract each my definition from excel and add to a list
for i in range(4, 74):
    cell_obj = sheet_obj.cell(row=i, column=9)
    myDef = cell_obj.value
    myDefinition.append(myDef)
    mydefLen = len(myDefinition)

# tests for matching lenghts
#testLenUrlvMyDef()

# for url in urls:
#     result = requests.get(url)
#     src = result.content
#     soup = BeautifulSoup(src, 'lxml')
#     while attempts < 2:
#         try:
#             name = soup.find('h2', {'class': ''}).next_element
#             # print(name)
#             productNames.append(name)
#         except AttributeError:
#             print("i encountered and error try again")
#             time.sleep(2)
#             attempts += 1
#             continue
#         else:
#             break
#     else:
#         productNames.append("None")
#         print(url)
#     time.sleep(5)
# print(productNames)
# print(len(productNames))

for url in urls:
    result = requests.get(url)
    src = result.content
    soup = BeautifulSoup(src, 'lxml')
    while attempts2 < 2:
        try:
            price = soup.find('h2', {'class': 'bop-price__current'})
            # print(price.text.strip())
            productPrices.append(price.text.strip())
        except AttributeError:
            print("i encountered and error try again")
            time.sleep(2)
            attempts2 += 1
            continue
        else:
            break
    else:
        productPrices.append(0)
        print(url)

    time.sleep(5)

print(productPrices)
print(len(productPrices))

# for url in urls:
#     result = requests.get(url)
#     src = result.content
#     soup = BeautifulSoup(src, 'lxml')
#     while attempts3 <2:
#         try:
#             amount = soup.find('span', {'class': 'bop-catchWeight'})
#             # print(amount.text.strip())
#             productAmounts.append(amount.text.strip())
#         except AttributeError:
#             print("i encountered and error try again")
#             time.sleep(2)
#             attempts3 += 1
#             continue
#         else:
#             break
#     else:
#         productAmounts.append(0)
#         print(url)
#     time.sleep(5)
# print(productAmounts)
# print(len(productAmounts))









